import openpyxl
from openpyxl.styles import PatternFill, colors, Font, Alignment

excel_file = 'data/JBFG_임직원.xlsx'

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# 1) 셀의 배경색상 지정!!!
pattern_red    = PatternFill(start_color='FF0000', fill_type='solid')
pattern_yellow = PatternFill(start_color='FFFF00', fill_type='solid')
# pattern_blue   = PatternFill(start_color='0000FF', fill_type='solid')
pattern_blue   = PatternFill(start_color=colors.BLUE, fill_type='solid')

ws.cell(1, 2).fill = pattern_red
ws.cell(4, 4).fill = pattern_yellow
ws.cell(5, 3).fill = pattern_blue


# 2) 폰트 지정
font_20 = Font(name = '맑은고딕', size = 20, color=colors.BLUE)
ws.cell(3,1).font = font_20
ws.cell(3,3).font = font_20


# 3) 정렬
align_center = Alignment(horizontal='center', vertical='center')
for idx in range(1,6):
    ws.cell(1,idx).alignment = align_center

for idx in range(1,7):
    ws.cell(idx,2).alignment = align_center

# 4) 셀크기 조정
ws.row_dimensions[1].height = 50

ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 25


wb.save(excel_file)

print('프로그램을 종료합니다!!!')

