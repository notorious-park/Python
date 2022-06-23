import openpyxl

excel_file = 'data/JBFG_임직원.xlsx'

wb = openpyxl.load_workbook(excel_file)
ws = wb.create_sheet('전북은행')   # 새로운 시트 생성!!!!!!!!!!!!!

# 데이터 추가 방법 1
ws['A1'] = '사번'
ws['B1'] = '이름'
ws['C1'] = '부서명'
ws['D1'] = '연락처'
ws['E1'] = '이메일'

# 데이터 추가 방법 3
ws.append(['201','한혜형','디지털영업팀', '010-9876-1111','aa@jbbank.com'])
ws.append(['202','김영목','카드사업부',  '010-9876-2222','bb@jbbank.com'])
ws.append(['203','박성실','데이터분석부', '010-9876-3333','cc@jbbank.com'])
ws.append(['204','박요온','데이터분석부', '010-9876-4444','dd@jbbank.com'])
ws.append(['205','오승현','지주',       '010-9876-5555','ee@jbbank.com'])
ws.append(['301','김진수','전략기획부',  '010-9876-6666','zz@jbbank.com'])

wb.save(excel_file)

print('프로그램 종료!!!')