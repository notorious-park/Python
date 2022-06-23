import openpyxl
from openpyxl.utils import get_column_letter   # 칼럼값을 숫자값으로 변경해주는 모듈

excel_file = 'data/JBFG_임직원.xlsx'

wb = openpyxl.load_workbook(excel_file)
ws = wb['전북은행']

# 데이터 삭제
# del ws['A7']

# row_no = 7
# col_nm = get_column_letter(2)   # A : 1   /   B : 2   / C : 3 ......
# print('{}{}'.format(col_nm, row_no))
#
# cell_loc = '{}{}'.format(col_nm, row_no)
# print('{} : {}'.format(cell_loc, ws[cell_loc].value))   # 변수값 확인 : ws[cell_loc].value




# 한줄의 데이터 삭제
row_no = 7
for col_no in range(1,6):
    col_nm = get_column_letter(col_no)
    print('{}{}'.format(col_nm,row_no))
    cell_loc = '{}{}'.format(col_nm,row_no)
    del ws[cell_loc]

wb.save(excel_file)

print('프로그램 종료!!!')