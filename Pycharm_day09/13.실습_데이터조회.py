import openpyxl
from openpyxl.utils import get_column_letter   # 칼럼값을 숫자값으로 변경해주는 모듈

excel_file = 'data/JBFG_임직원.xlsx'

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# 방법1) 행열의 갯수를 알고 있을때
for row_no in range(1,7):
    for col_no in range(1,6):
        cell_loc = ws.cell(column=col_no, row=row_no).value
        print(cell_loc,end= '\t')
    print()

print('-'*100)

# 방법2) 행열의 갯수를 모를때
rows = ws.rows
# print('rows : {}, {}'.format(type(rows),rows))

arr_row = list()
cnt = 0
for row in rows:
    arr_row.append(row)
    for cell in row:
        val = cell.value
        print(val, end = '\t')
    print()
