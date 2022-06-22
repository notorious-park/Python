import openpyxl

excel_file = 'data/JBFG_임직원.xlsx'

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# 데이터 조회
A1 = ws['A1'].value
print('A1 = {}'.format(A1))

B1 = ws.cell(column=2, row=1).value
print('B1 = {}'.format(B1))

B2 = ws.cell(column=2, row=2).value
print('B2 = {}'.format(B2))

wb.close()   # 조회다보니 별도 저장 안해도됨 / 다만, close는 해줄 것!

print('프로그램 종료!!!')