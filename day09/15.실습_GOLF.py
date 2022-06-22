import openpyxl
from openpyxl.utils import get_column_letter
from random import randint
from openpyxl.styles import PatternFill, colors, Font, Alignment

excel_file = "data/golf_score_board.xlsx"

wb = openpyxl.load_workbook(excel_file)

ws = wb['스코어']

# 미션1 : 참가선수들 등록 (실시간입력/등록명단파일호출)

def 선수등록(filename):
    # step1   :  file 불러오기
    file = open("{}".format(filename), "r", encoding='utf-8')
    line = file.readlines()
    file.close()

    # step2 : 참가선수를 리스트형태로 변환
    peoples = []
    for idx in range(len(line)):
        people = line[idx].replace('\n', '')  # 참가선수1명씩 가져옴.
        peoples.append(people)  # 전체참가선수를 리스트형태로 저장

    for idx in range(len(peoples)):  # 엑셀에 참가선수 뿌리기
        b = ws.cell(row=idx + 3, column=1, value=peoples[idx])
        print(b.value)
    return peoples
    print('참가선수 등록 완료')

player = 선수등록("./data/명단.txt")

# 미션2 : 참가선수들 HOLE별 스코어를 랜덤하게 입력

def 난수뿌리기():
    for row_no in range(3,len(player)+3):
        for col_no in range(3, 21):
            inp_nm = '{}{}'.format(get_column_letter(col_no), 2)
            par_ran = randint(1, ws[inp_nm].value*2)
            score = par_ran - ws[inp_nm].value
            # ws.cell(column = col_no, row = row_no, value = par_ran)
            ws.cell(column = col_no, row = row_no, value = score)
        # return score
난수뿌리기()

# # 미션3 : 18홀까지의 스코어, 보정치를 계산하여 등록 (핸디는 옵션), 보정치는 없다.

def 스트로크합산():
    for col_no in range(3, len(player)+3):
        arr_raw = list()
        for a in range(3, 21):
            ran  = '{}{}'.format(get_column_letter(a), col_no)
            arr_raw.append(ws[ran].value)
        stroke = sum(arr_raw)+72
        ws.cell(column = 21, row = col_no, value = stroke)
        ws.cell(column=23, row=col_no, value = stroke)
스트로크합산()

# # 미션4 : 보정치 기준으로 랭킹 등록, 각종 기록 갯수 등록

strokes = []

for i in range(len(player)):
    strokes.append(ws.cell(column=23, row=i + 3).value)  # 점수값 가져오기
players_stroke=strokes

def 선수점수정렬():       #높은 순서대로 정렬
    sort_score = sorted(players_stroke)
    return sort_score

def 등수만들기():
    rank_dict = {}  #{ 100점 : 1등, 90점 : 2등 ... } 로 구성된 딕셔너리
    rank = 1        #1등부터 시작
    sort_score = 선수점수정렬()
    for i in range(len(sort_score)):
        rank_dict[sort_score[i]] = rank  #점수별로 등수 매겨준다.
        if sort_score[i] == sort_score[i-1]:  #동점이라면?
            rank_dict[sort_score[i]] = rank - sort_score.count(sort_score[i]) + 1  # (예 : 90,90,80,79,60 -> 1,1,3,4,5 로 등수 나옴 )
        else:
            pass
        rank += 1
    return rank_dict

def 랭킹채우기():
    #player = mp()
    rank_dict = 등수만들기()
    for row_idx in range(3, len(player) +3):
        score = ws.cell(column = 23, row = row_idx ).value  #21-> 23으로 바꿔야함 -> 랭킹(X)컬럼을 가져온다.
        if score in rank_dict:
            rank = rank_dict[score] #등수가져옴(딕셔너리의 value를 추출)
            ws.cell(column= 24, row=row_idx, value=rank) #삽입위치
        else:
            pass
랭킹채우기()

# # 이글 등 카운트 없으면 0으로 할지
font_color = Font(color=colors.WHITE) #양파 추가!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
def 기록카운트():
    for row_no in range(3,len(player)+3):
        eagle  = []
        birdie = []
        par    = []
        bogey  = []
        double_pa = []
        for col_no in range(3, 21):
            idx = ws.cell(column=col_no, row=row_no).value
            if ws.cell(column = col_no, row=2 ).value == ws.cell(column = col_no, row=row_no ).value:
                double_pa.append(1)
                sum_double_pa = sum(double_pa)
                ws.cell(column=29, row=row_no, value=sum_double_pa)
                ws.cell(column=29, row=row_no).font = font_color
            elif idx == -2:
                eagle.append(1)
                sum_eagle = sum(eagle)
                ws.cell(column=25, row=row_no, value=sum_eagle)
            elif idx == -1:
                birdie.append(1)
                sum_birdie = sum(birdie)
                ws.cell(column=26, row=row_no, value=sum_birdie)
            elif idx == 0:
                par.append(1)
                sum_par = sum(par)
                ws.cell(column=27, row=row_no, value=sum_par)
            elif idx == 1:
                bogey.append(1)
                sum_bogey = sum(bogey)
                ws.cell(column=28, row=row_no, value=sum_bogey)
            else:
                pass
기록카운트()

###화요일###
# # 미션5 : 대회결과 시트에 결과에 맞는 선수명과 최종성적/기록수를 등록

ws2 = wb['대회결과']

def 대회결과():
    #sheet2번 첫번째 표 채우기 (대회결과)
    infos = 선수정보담기()
    sorted_infos = sorted(infos, key = lambda x: (x['스트로크']))  #스트로크을 가지고 정렬한다.  -> 최종성적으로 바꾸기

    row_idx = 2  #2행부터 채울거니까
    for i in [0,1,2,-1]:  #앞에 상위 3명, 뒤에 하위 1명 가져온다.
        info = sorted_infos[i]
        ws2.cell(column=1, row=row_idx, value='꼴등' if i == -1 else str(info['랭킹']) + '등')  #A컬럼채우기
        ws2.cell(column=2, row=row_idx, value=info['선수명'])                                  #B컬럼채우기
        ws2.cell(column=3, row=row_idx, value=info['스트로크'])                                #C컬럼채우기

        row_idx += 1  #2,3,4,5행에 넣을 거니까
대회결과()


def 선수정보담기():
    #sheet2번에서 활용할 용도  (선수
    infos = list()
    for row_idx in range(3, len(player) +3):
        person_dict = {}
        person_dict['선수명'] = ws.cell(column = 1, row=row_idx).value
        person_dict['스트로크'] = ws.cell(column=21, row=row_idx).value
        person_dict['랭킹']  =  ws.cell(column = 24, row=row_idx).value
        person_dict['버디']  =  ws.cell(column = 26, row=row_idx).value
        person_dict['파'] = ws.cell(column=27, row=row_idx).value
        person_dict['보기'] = ws.cell(column=28, row=row_idx).value
        person_dict['양파'] = ws.cell(column=29, row=row_idx).value

        infos.append(person_dict)
    return infos
선수정보담기()

def 버디최고기록(infos):

    many_birdie = sorted(infos, key=lambda x: (x['버디']),reverse=True)[0]   #버디많은 사람 순서대로 정렬했을때 첫번째 사람을 가져온다.
    ws2.cell(column=2, row=8, value=many_birdie['선수명'])
    ws2.cell(column=3, row=8, value=many_birdie['버디'])

def 파최고기록(infos):
    many_par = sorted(infos, key=lambda x: (x['파']),reverse=True)[0]      #파많은 사람 순서대로 정렬했을때 첫번째 사람을 가져온다.
    ws2.cell(column=2, row=9, value=many_par['선수명'])
    ws2.cell(column=3, row=9, value=many_par['파'])

def 보기최고기록(infos):
    many_bogey = sorted(infos, key=lambda x: (x['보기']),reverse=True)[0]   #보기많은 사람 순서대로 정렬했을때 첫번째 사람을 가져온다.
    ws2.cell(column=2, row=10, value=many_bogey['선수명'])
    ws2.cell(column=3, row=10, value=many_bogey['보기'])

def 양파최고기록(infos): #양파 추가!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    many_bogey = sorted(infos, key=lambda x: (x['양파']),reverse=True)[0]   #양파많은 사람 순서대로 정렬했을때 첫번째 사람을 가져온다.
    ws2.cell(column=2, row=11, value=many_bogey['선수명'])
    ws2.cell(column=3, row=11, value=many_bogey['양파'])


def 최고기록선수넣기():
    infos = 선수정보담기()
    버디최고기록(infos)
    파최고기록(infos)
    보기최고기록(infos)
    양파최고기록(infos) #양파 추가!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!

최고기록선수넣기()





wb.save(excel_file)
print('프로그램 종료')