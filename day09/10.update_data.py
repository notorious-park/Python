import openpyxl

excel_file = 'data/JBFG_임직원.xlsx'

wb = openpyxl.load_workbook(excel_file)

# 시트 지정
ws = wb['전북은행']

# 기존 데이터 수정
ws['C7'] = '빅파이크래프트'
ws['D7'] = '010-9988-1122'
ws['E7'] = 'bigpycraft@gmail.com'


wb.save(excel_file)

print('프로그램 종료!!!')