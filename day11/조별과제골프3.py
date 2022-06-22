
import openpyxl
from openpyxl.utils import get_column_letter
from random import randint
from openpyxl.styles import PatternFill, colors, Font, Alignment

excel_file = "data/golf_score_board.xlsx"

wb = openpyxl.load_workbook(excel_file)

ws = wb['스코어']

# 미션1 : 참가선수들 등록 (실시간입력/등록명단파일호출)

def mp(filename):
    #step1   :  file 불러오기
    file = open("{}".format(filename), "r",encoding = 'utf-8')
    line = file.readlines()
    file.close()

    #step2 : 참가선수를 리스트형태로 변환
    peoples  = []
    for idx in range(len(line)):
        people = line[idx].replace('\n','')  #참가선수1명씩 가져옴.
        peoples.append(people)               #전체참가선수를 리스트형태로 저장

    for idx in range(len(peoples)):                 #엑셀에 참가선수 뿌리기
        b = ws.cell(row=idx+3,column=1,value=peoples[idx])
        print(b.value)
    return peoples
    print('참가선수 등록 완료')

"""
    1단계: 선수등록
        - 선수명단 파일로 저장 : csv
        - 선수정보로딩 : CSV파일 READ
        - 선수등록 : 엑셀파일에 저장
"""
player = mp("./data/명단.txt")

## 교수님 작성 ##

# def 선수정보로딩():
#     player = list()
#     pass
#     return True, player
#
# def 선수등록(player):
#     pass
#     return True
#
#
# def step1_선수등록():
#
#     res, player = 선수정보로딩()
#
#     if res:
#         선수등록(player)
#
# mp2("./data/명단.txt")

# 미션2 : 참가선수들 HOLE별 스코어를 랜덤하게 입력

def play_game():
    for row_no in range(3,len(player)+3):
        for col_no in range(3, 21):
            inp_nm = '{}{}'.format(get_column_letter(col_no), 2)
            par_ran = randint(1, ws[inp_nm].value*2)
            score = par_ran - ws[inp_nm].value
            # ws.cell(column = col_no, row = row_no, value = par_ran)
            ws.cell(column = col_no, row = row_no, value = score)
        # return score
play_game()


###월요일###
# # 미션3 : 18홀까지의 스코어, 보정치를 계산하여 등록 (핸디는 옵션), 보정치는 없다.
# # 스트로크 = 최종성적
# 영목, 요온코드
def stroke_sc():
    for col_no in range(3, len(player)+3):
        arr_raw = list()
        for a in range(3, 21):
            ran  = '{}{}'.format(get_column_letter(a), col_no)
            arr_raw.append(ws[ran].value)
        stroke = sum(arr_raw)+72
        ws.cell(column = 21, row = col_no, value = stroke)
stroke_sc()

'''은민대리님 코드
def sum_game():
    """
    each_score :  각 선수들의 par1~par18까지의 점수합
    player_score : 모든 선수들의 접수합을 리스트로 묶어둠
    player_names : 모든 선수들
    :return:
    """
    player_score = []
    player_names = []
    for row_idx in range(3,len(player)+3):
        each_score = 0
        for col_idx in range(3,21):
            cell_val = ws.cell(column=col_idx, row = row_idx).value
            each_score += cell_val
            player_name = ws.cell(column=1, row = row_idx).value
        ws.cell(column = 21, row = row_idx, value = each_score)        # U열에 넣기

        player_score.append(each_score)
        player_names.append(player_name)
    return player_names, player_score

sum_game()  #함수실행
player_names, player_score = sum_game()  #sum_game() 함수의 리턴값을 받는다.  -> 미션4에서 활용
'''

# # 미션4 : 보정치 기준으로 랭킹 등록, 각종 기록 갯수 등록
# # 랭킹
# def ranking():
#     arr_raw = list()
#     arr_raw2 = list()
#     for col_no in range(3, len(player)+3):
#         stroke = '{}{}'.format('U', col_no)
#         who = '{}{}'.format('A', col_no)
#         arr_raw = '{},{}'.format(ws[who].value, ws[stroke].value)
#         arr_raw2.append(arr_raw)
#     print(arr_raw2)
#     print(type(arr_raw2))
#         # print(arr_raw.values())
#     # for i in range(arr_raw):
#     #     print(arr_raw)
#     #     print(len(arr_raw))
# ranking()

###화요일###
# # 미션5 : 대회결과 시트에 결과에 맞는 선수명과 최종성적/기록수를 등록

wb.save(excel_file)
print('프로그램 종료')
