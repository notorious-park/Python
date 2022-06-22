import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from random import randint
from openpyxl.styles import PatternFill, colors, Font, Alignment

excel_file = "data/golf_score_board.xlsx"

wb = openpyxl.load_workbook(excel_file)

ws = wb['스코어']

# 미션1 : 참가선수들 등록 (실시간입력/등록명단파일호출)

def mp(filename):
    #step1   :  file 불러오기
    file = open("{}".format(filename), "r",encoding = 'utf-8')
    line = file.readlines()
    file.close()

    #step2 : 참가선수를 리스트형태로 변환
    peoples  = []
    for idx in range(len(line)):
        people = line[idx].replace('\n','')  #참가선수1명씩 가져옴.
        peoples.append(people)               #전체참가선수를 리스트형태로 저장

    for idx in range(len(peoples)):                 #엑셀에 참가선수 뿌리기
        b = ws.cell(row=idx+3,column=1,value=peoples[idx])
        print(b.value)
    return peoples
    print('참가선수 등록 완료')

"""
    1단계: 선수등록
        - 선수명단 파일로 저장 : csv
        - 선수정보로딩 : CSV파일 READ
        - 선수등록 : 엑셀파일에 저장
"""
player = mp("./data/명단.txt")

## 교수님 작성 ##

# def 선수정보로딩():
#     player = list()
#     pass
#     return True, player
#
# def 선수등록(player):
#     pass
#     return True
#
#
# def step1_선수등록():
#
#     res, player = 선수정보로딩()
#
#     if res:
#         선수등록(player)
#
# mp2("./data/명단.txt")

# 미션2 : 참가선수들 HOLE별 스코어를 랜덤하게 입력

def play_game():
    for row_no in range(3,len(player)+3):
        for col_no in range(3, 21):
            inp_nm = '{}{}'.format(get_column_letter(col_no), 2)
            par_ran = randint(-2, ws[inp_nm].value)
            # print(par_ran)
            ws.cell(column = col_no, row = row_no, value = par_ran)
play_game()

###월요일###
# 미션3 : 18홀까지의 스코어, 보정치를 계산하여 등록 (핸디는 옵션), 보정치는 없다.

def sum_scr():
    for row_no in range(3,len(player)+3):
        scr_value = []
        for col_no in range(3, 21):
            idx = ws.cell(column=col_no, row=row_no).value
            scr_value.append(idx)
            sum_score = sum(scr_value)
            ws.cell(column = 21, row = row_no, value = sum_score + 72)
sum_scr()

# 미션4 : 보정치 기준으로 랭킹 등록, 각종 기록 갯수 등록
def par_dvcd():
    for row_no in range(3,len(player)+3):
        eagle  = []
        birdie = []
        par    = []
        bogey  = []
        for col_no in range(3, 21):
            idx = ws.cell(column=col_no, row=row_no).value
            if idx == -2:
                eagle.append(1)
                sum_eagle = sum(eagle)
                ws.cell(column=25, row=row_no, value=sum_eagle)
            elif idx == -1:
                birdie.append(1)
                sum_birdie = sum(birdie)
                ws.cell(column=26, row=row_no, value=sum_birdie)
            elif idx == 0:
                par.append(1)
                sum_par = sum(par)
                ws.cell(column=27, row=row_no, value=sum_par)
            elif idx == 1:
                bogey.append(1)
                sum_bogey = sum(bogey)
                ws.cell(column=28, row=row_no, value=sum_bogey)
            else:
                pass
par_dvcd()




###화요일###
# # 미션5 : 대회결과 시트에 결과에 맞는 선수명과 최종성적/기록수를 등록

wb.save(excel_file)
print('프로그램 종료')
